'''
DAILY REPORT AUTOMATION RECORD KEEPING:
This records transactions on a daily basis by account number that can write directly to the file.
The  import openpyxl as xl  is the ONLY Python way to directly write to an Excel file. 
This is important because using a dataframe and resaving a new file will cause the formulas and formating to disaper. 
'''

import openpyxl as xl
import pandas as pd

path1 = r'C:\Users\scott\OneDrive\Desktop\Excel\formulas.xlsx'
path2 = r'C:\Users\scott\OneDrive\Desktop\Excel\formulas1.xlsx'

wb1 = xl.load_workbook(filename=path1)
ws1 = wb1.worksheets[0]

wb2 = xl.load_workbook(filename=path2)
ws2 = wb2.create_sheet(ws1.title)

# give column name
col_name = "July"
  
# find the index no

df = pd.read_excel(path1)
#df.columns=df.iloc[0]
index_no = df.columns.get_loc(col_name)
abc={1:'A', 2:'B', 3:'C', 4:'D', 5:'E', 6:'F', 7:'G', 8:'H', 9:'I'}
letter=abc[index_no] #transfers index no to Excel Col letter

amount=[4, 8, 12, 7, 10, 23, 6, 8]


for row in ws1:
    for cell in row:
        #print(cell.coordinate) #A1  B1  C1   #type is 'str'
        ws2[cell.coordinate].value = cell.value

for a in amount:
    for i in range(2, len(amount)):
        cc=letter+str(i)
        ws2[cc].value = a
        
wb2.save(path2)
wb1.close()
wb2.close()
print("Done")
